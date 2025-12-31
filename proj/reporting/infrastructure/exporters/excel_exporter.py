"""Excel exporter for reporting.

Uses openpyxl for Excel file generation with formatting (bold headers, color-coded status).
Falls back to CSV if openpyxl is not available.
"""
from typing import Iterable, Mapping, Any
import io


def write_excel(path: str, rows: Iterable[Mapping[str, object]]) -> None:
    """Write rows to an Excel file at `path` with formatting.
    
    Formatting includes:
    - Bold header row
    - Green cells for "Present" status
    - Red cells for "Absent" status
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        from .csv_exporter import write_csv
        csv_path = path.replace('.xlsx', '.csv') if path.endswith('.xlsx') else path
        write_csv(csv_path, rows)
        return

    from .csv_exporter import CSV_HEADER

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Write header with bold font
    bold_font = Font(bold=True)
    for col_num, header in enumerate(CSV_HEADER, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

    # Color fills for status
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    # Write data rows with conditional formatting
    for row_num, row_data in enumerate(rows, start=2):
        status_col = None
        for col_num, header in enumerate(CSV_HEADER, start=1):
            value = row_data.get(header) if isinstance(row_data, dict) else getattr(row_data, header, "")
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Track status column for formatting
            if header == "status":
                status_col = col_num
                if value == "Present":
                    cell.fill = green_fill
                elif value == "Absent":
                    cell.fill = red_fill

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(path)


class ExcelExporter:
    """Excel exporter using openpyxl."""

    def export_bytes(self, payload: Mapping[str, Any]) -> bytes:
        """Produce Excel bytes for the given payload.
        
        Expected payload shape: {"students": [ {..row..}, ... ], ...}
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV bytes if openpyxl not available
            from .factory import CsvExporter
            return CsvExporter().export_bytes(payload)

        from .csv_exporter import CSV_HEADER

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Write header with bold font
        bold_font = Font(bold=True)
        for col_num, header in enumerate(CSV_HEADER, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font

        # Color fills for status
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # Write data rows
        rows = payload.get("students", []) or []
        for row_num, row_data in enumerate(rows, start=2):
            for col_num, header in enumerate(CSV_HEADER, start=1):
                value = row_data.get(header) if isinstance(row_data, dict) else getattr(row_data, header, "")
                cell = ws.cell(row=row_num, column=col_num, value=value)
                
                # Apply conditional formatting for status column
                if header == "status":
                    if value == "Present":
                        cell.fill = green_fill
                    elif value == "Absent":
                        cell.fill = red_fill

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
